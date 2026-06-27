# backend/exports/excel_export.py
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List

def generate_excel(df: pd.DataFrame, kpis: List[Dict[str, Any]], stats_results: Dict[str, Any]) -> io.BytesIO:
    wb = Workbook()
    
    # Fonts & Fills
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=14, bold=True, color="0F172A")
    normal_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep navy
    kpi_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Light gray
    
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light green
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light red
    orange_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Light orange
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # ----------------------------------------------------
    # Sheet 1: Raw Cleaned Data
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Cleaned Data"
    
    # Write Title
    ws1.cell(row=1, column=1, value="Boardroom-AI - Cleaned Dataset Report").font = title_font
    
    # Headers
    headers = list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Data Rows
    for row_idx, row in enumerate(df.values, 4):
        for col_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=str(val) if pd.notnull(val) else "")
            cell.font = normal_font
            cell.border = thin_border
            
            # Format alignment
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
                
    # ----------------------------------------------------
    # Sheet 2: KPI Summary Table
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="KPI Dashboard")
    ws2.cell(row=1, column=1, value="Key Performance Indicators (KPIs)").font = title_font
    
    kpi_headers = ["KPI Name", "Calculated Value", "Source Column", "Trend Direction", "Brief Insight"]
    for col_idx, h in enumerate(kpi_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    for idx, kpi in enumerate(kpis, 4):
        ws2.cell(row=idx, column=1, value=kpi.get("name")).font = bold_font
        
        val_cell = ws2.cell(row=idx, column=2, value=kpi.get("value"))
        val_cell.font = bold_font
        
        # Color coding base on trend
        trend = kpi.get("trend", "flat")
        if trend == "up":
            val_cell.fill = green_fill
        elif trend == "down":
            val_cell.fill = red_fill
        else:
            val_cell.fill = kpi_fill
            
        ws2.cell(row=idx, column=3, value=kpi.get("column")).font = normal_font
        ws2.cell(row=idx, column=4, value=kpi.get("trend").upper()).font = normal_font
        ws2.cell(row=idx, column=5, value=kpi.get("insight")).font = normal_font
        
        for c in range(1, 6):
            ws2.cell(row=idx, column=c).border = thin_border
            
    # ----------------------------------------------------
    # Sheet 3: Statistical Analysis Results
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Statistical Analysis")
    ws3.cell(row=1, column=1, value="Descriptive Statistics").font = title_font
    
    stat_headers = ["Metric/Column", "Mean", "Median", "Mode", "Std Dev", "Min", "Max", "Distribution Shape"]
    for col_idx, h in enumerate(stat_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    row_counter = 4
    for col_name, col_stats in stats_results.get("descriptive_stats", {}).items():
        ws3.cell(row=row_counter, column=1, value=col_name).font = bold_font
        ws3.cell(row=row_counter, column=2, value=col_stats.get("mean")).font = normal_font
        ws3.cell(row=row_counter, column=3, value=col_stats.get("median")).font = normal_font
        ws3.cell(row=row_counter, column=4, value=col_stats.get("mode")).font = normal_font
        ws3.cell(row=row_counter, column=5, value=col_stats.get("std")).font = normal_font
        ws3.cell(row=row_counter, column=6, value=col_stats.get("min")).font = normal_font
        ws3.cell(row=row_counter, column=7, value=col_stats.get("max")).font = normal_font
        
        dist_cell = ws3.cell(row=row_counter, column=8, value=col_stats.get("distribution"))
        dist_cell.font = normal_font
        
        # Color coding distribution
        dist = col_stats.get("distribution", "Normal")
        if dist == "Normal":
            dist_cell.fill = green_fill
        elif "Skewed" in dist:
            dist_cell.fill = orange_fill
        else:
            dist_cell.fill = kpi_fill
            
        for c in range(1, 9):
            cell = ws3.cell(row=row_counter, column=c)
            cell.border = thin_border
            if c > 1 and c < 8:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
        row_counter += 1
        
    # Auto-fit column widths for all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save workbook to memory
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file
